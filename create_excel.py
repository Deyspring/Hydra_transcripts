#! python
# Create_excel makes an excel file with a title based on a website and the date. 
###TODO make sure this always saves a worksheet with a different name so they don't get saved over each other.

import openpyxl
from openpyxl.styles import Font
from datetime import datetime


def create_header(): 
	'''Excel file with a title based on a website and the date'''

	#TODO? make these names and link changable in a gui? 
	title = 'Pfam Domains'   
	subtitle = 'Pfam domains in predicted Hydra proteins'
	web_address = 'https://research.nhgri.nih.gov/hydra/pfam/'


	title = title.replace(" ","")
	wb = openpyxl.Workbook() # Create a blank workbook
	wb.sheetnames # It starts with one sheet
	sheet = wb.active
	sheet.title
	sheet.title = title # Change titles

	#Month abbreviation, day and year
	now = datetime.now()
	datestamp = now.strftime("%b_%d_%Y_%H_%M")

	#Create fonts
	title_style = Font(size=24, bold=True) 

	#Format spreadsheet
	sheet['A1'] = title
	sheet['A1'].font = title_style
	sheet['A2'] = subtitle
	sheet['A3'] = web_address
	sheet['A5'] = datestamp 
	wb.save( title+ '_' + datestamp + '.xlsx' ) # Save the workbook.
	file_name = (title+ '_' + datestamp + '.xlsx')
	print('saved workbook')
	return file_name

#create_header()

def add_data_xl(search_term, geneids, transcripts, file_name): 
	'''Update excel file with a search term subheader'''

	wb = openpyxl.load_workbook(file_name)
	sheet = wb.active
	# Start of search term, geneid and transcripts area. 
	row = 11 #first row after header
	column = 1 

	# Search term header for each search term 
	search_term = sheet.cell(row, column).value
	
	for geneid in geneids: 
		row += 1
		geneid_cell = sheet.cell(row, column).value 

		for transcript in transcripts: 
			column +=1 
			transcript_cell = sheet.cell(row, column).value 

	wb.save(file_name)


term = 'ig'
geneids = ['badger','badger2','mushroom','mushroom2']
transcripts = [['bad','ger','bad','ger'],['bad2','ger2','bad2','ger2'],['snake','oh','snake'],['mush','room','mush','room']]
file_name ='PfamDomains_Feb_04_2020_20_07.xlsx'

add_data_xl(term, geneids,transcripts,file_name)


"""
def add_term_xl(search_term): 
	'''Update excel file with a search term and it's resulting geneids and transcripts'''
	# Search term 
	search_term_header = sheet.cell(row = 11, column =1).value
	#Make a gene identifier as a test case 
	geneid_cell = sheet.cell( row = 12, column = 1 ).value

	#and the transcripts associated with each gene
	transcript_cell = ['a','b','c']

	# Todo: loop through all the rows and add the gene identifiers and transcripts 

	for rowNum in range(11, sheet.max_row):  # skip the first rows
		gene_Ident = sheet.cell(row=rowNum, column=1).value
	
		colNum = 2

	for i in transcripts: 
		sheet.cell(row=rowNum, column=colNum).value = i
		colNum += 1

	print (sheet.cell(row=11, column=2).value)

	wb.save('UpdateHydra.xlsx')"""

