#! python
# Create_excel makes an excel file with a title based on a website and the date. 

import openpyxl
from openpyxl.styles import Font
import time

#if __name__ == '__main__': #not sure what to name instead of _main_ for clarity 
 
# main()

def mng_xl(terms, geneid_lists, transcript_sets): 
	'''Excel file with a title based on a website and the date'''

	#TODO? make these names and link changable in a gui? 
	title = 'Pfam_Domains'   
	subtitle = 'Pfam domains in predicted Hydra proteins'
	web_address = 'https://research.nhgri.nih.gov/hydra/pfam/'
	datestamp = time.strftime("%Y-%m-%d %H:%M")

	title = title.replace(' ','')
	wb = openpyxl.Workbook() # Create a blank workbook
	wb.sheetnames # It starts with one sheet
	sheet = wb.active
	sheet.title
	sheet.title = title 

	#Create fonts
	title_style = Font(size=24, bold=True) 
	term_style = Font(size =16, bold = True)
	term_sub_style = Font(size =16)
	geneid_style = Font(size=14, bold = True)
	transcript_style = Font(size=12, bold = True)

	#Format spreadsheet
	sheet['A1'] = title
	sheet['A1'].font = title_style
	sheet['A2'] = subtitle
	sheet['A3'] = web_address
	sheet['A5'] = datestamp 
	blank_cell= " "
	
	row_num = 6
	col_num = 1

	row_num += 2
	sheet.cell(row=row_num, column=col_num).font = term_style
	sheet.cell(row=row_num, column=col_num).value = "TERM"

	print("manage_excel\n")


	if len(terms) <= 1:
		print("one term:",terms)
		term = str(terms)

		row_num += 1
		sheet.cell(row=row_num, column=col_num).font = term_sub_style
		sheet.cell(row=row_num, column=col_num).value = term

		row_num += 1
		sheet.cell(row=row_num, column=col_num).font = geneid_style
		sheet.cell(row=row_num, column=col_num).value = "Geneids"
		col_num +=1
		sheet.cell(row=row_num, column=col_num).font = transcript_style
		sheet.cell(row=row_num, column=col_num).value = "transcripts"
		col_num -=1 

		for geneid, transcripts in zip(geneid_lists, transcript_sets):
				print("geneid: ", geneid)
				row_num += 1
				col_num = 1
				sheet.cell(row=row_num, column=col_num).value = geneid
			
				for transcript in transcripts: 
					print("transcript:", transcript)
					col_num +=1
					sheet.cell(row=row_num, column=col_num).value = transcript

	else: 

		for term, geneid_list, transcript_set in zip(terms, geneid_lists, transcript_sets): 
			# Search terms should start at cell A7, row 
			print("multiple terms:", term )
			print(type(term))

			col_num = 1
			row_num += 1
			sheet.cell(row=row_num, column=col_num).font = term_sub_style
			sheet.cell(row=row_num, column=col_num).value = term

			row_num += 1
			sheet.cell(row=row_num, column=col_num).font = geneid_style
			sheet.cell(row=row_num, column=col_num).value = "Geneids"
			col_num +=1
			sheet.cell(row=row_num, column=col_num).font = transcript_style
			sheet.cell(row=row_num, column=col_num).value = "transcripts"
			col_num -=1 

			for geneid, transcripts in zip(geneid_list, transcript_set):
				print("geneid: ", geneid)
				row_num += 1
				col_num = 1
				sheet.cell(row=row_num, column=col_num).value = geneid
			
				for transcript in transcripts: 
					print("transcript:", transcript)
					col_num +=1
					sheet.cell(row=row_num, column=col_num).value = transcript

	

	wb.save( title+'_'+ datestamp+'.xlsx' ) # Save the workbook.
	file_name = (title+'_'+datestamp+'.xlsx')

	print('saved workbook\n')

#Test data 
"""
terms = ['Ig_4'] 
geneids = ['Sc4wPfr_172.1.g1567.t1', 'Sc4wPfr_435.1.g6986.t1', 'Sc4wPfr_274.1.g13601.t1', 'Sc4wPfr_172.2.g23178.t1', 'Sc4wPfr_268.g25749.t1', 'Sc4wPfr_268.g25754.t2', 'Sc4wPfr_268.g25749.t2', 'Sc4wPfr_268.g25754.t1'] 
transcripts = [{'t18444aep', 't20351aep', 't1128aep', 't19864aep', 't34681aep' }, 
			   {'t36740aep', 't12260aep', 't30006aep', 't37984aep'}, 
			   {'t30122aep', 't297aep'}, 
			   {'t18444aep', 't20351aep', 't1128aep', 't19864aep', 't9044aep', 't22011aep'}, 
			   {'t24725aep', 't27640aep', 't2308aep', 't31900aep', 't21629aep', 't15944aep', 't14058aep', 't21081aep', 't1197aep'}, 
			   {'t24725aep', 't27640aep', 't2308aep', 't31900aep', 't21629aep', 't12602aep', 't3984aep', 't27276aep', 't4494aep', 't13463aep', 't18097aep', 't36467aep', 't18096aep', 't23522aep', 't15944aep', 't26172aep', 't24866aep', 't20881aep', 't15548aep', 't439aep', 't14058aep', 't21081aep', 't1197aep'}, 
			   {'t24725aep', 't27276aep', 't15944aep', 't26172aep', 't24866aep', 't20881aep', 't15548aep', 't439aep', 't14058aep', 't21081aep', 't1197aep'}, 
			   {'t24725aep', 't27640aep', 't2308aep', 't31900aep', 't21629aep', 't12602aep', 't14827aep', 't23311aep', 't3984aep', 't27276aep', 't4494aep', 't13463aep', 't18097aep', 't36467aep', 't18096aep', 't23522aep', 't15944aep', 't26172aep', 't24866aep', 't20881aep', 't15548aep', 't439aep', 't14058aep', 't21081aep', 't1197aep'}]
"""
"""
terms = ['term1','term2']
geneids = ['1_1','1_2','1_3','1_4'],['2_1','2_2','2_3','2_4']
transcripts = [{'111','112','1113','1114'},{'2111','2112','2113','2114'},{'3111','3112','3114'},{'4111','4112','41113','4114'}],[{'000000000','term2_gene1_tr2','term2_gene1_tr3','term2_gene1_tr4'},{'##############','term2_gene2_tr2','term2_gene2_tr3','term2_gene2_tr4'},{'MMMMMMMMMMMMM','term2_gene3_tr2','term2_gene3_tr3'},{'term2_gene4_tr1','term2_gene4_tr2','term2_gene4_tr3','term2_gene4_tr'}]


mng_xl(terms,geneids,transcripts)

"""