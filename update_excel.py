#python!
#Hello? module description needed

pass #stub for testing

"""
import openpyxl



def update_xl(): 
	'''Update excel file with geneIDs and their transcripts'''

	#Make a gene identifier as a test case 
	gene_Ident = sheet.cell( row = 11, column = 1 ).value

	#and the transcripts associated with each gene
	transcripts = ['a','b','c']

	# Todo: loop through all the rows and add the gene identifiers and transcripts 

	for rowNum in range(11, sheet.max_row):  # skip the first rows
		gene_Ident = sheet.cell(row=rowNum, column=1).value
	
		colNum = 2

	for i in transcripts: 
		sheet.cell(row=rowNum, column=colNum).value = i
		colNum += 1

	print (sheet.cell(row=11, column=2).value)

	wb.save('UpdateHydra.xlsx')

========================================
	import openpyxl
import time

title = 'Pfam Domains'
subtitle = 'Pfam domains in predicted Hydra proteins'

def create_xl(title,subtitle):

#Create an excel file
wb = openpyxl.load_workbook('hydra.xlsx')
sheet = wb.get_sheet_by_name('Sheet1') #top sheet



wb.save('UpdateHydra.xlsx')


title = 'Pfam Domains'
subtitle = 'Pfam domains in predicted Hydra proteins'

def create_xl(title,subtitle):

#Create an excel file
wb = openpyxl.load_workbook('hydra.xlsx')
sheet = wb.get_sheet_by_name('Sheet1') #top sheet



wb.save('UpdateHydra.xlsx')




"""



